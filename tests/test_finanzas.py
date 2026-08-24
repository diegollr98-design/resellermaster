"""Tests de `core/finanzas.py` (superficie PERSISTENCIA: dinero).

Cubre: conversión céntimos<->euros, los agregados (total vendido, beneficio
bruto -- una 'devuelta' NO suma), y el export a `.xlsx` releído con openpyxl
(columnas + filas + el nombre por defecto con la fecha).
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pytest

from core.finanzas import (
    beneficio_total_cents,
    cents_a_euros,
    euros_texto_a_cents,
    exportar_excel,
    nombre_export_por_defecto,
    total_vendido_cents,
)

_COLUMNAS_ESPERADAS = (
    "Referencia",
    "Título",
    "Lote",
    "Coste (€)",
    "Subido",
    "Vendido",
    "Precio venta (€)",
    "Plataforma venta",
    "Fecha venta",
    "Beneficio bruto (€)",
)


# --------------------------------------------------------------------------
# cents_a_euros / euros_texto_a_cents
# --------------------------------------------------------------------------


def test_cents_a_euros_formatea_dos_decimales():
    assert cents_a_euros(1999) == "19.99"
    assert cents_a_euros(0) == "0.00"
    assert cents_a_euros(5) == "0.05"


def test_cents_a_euros_none_es_cadena_vacia():
    # None != 0: un importe NO INFORMADO no puede parecer "0.00" (un dato real).
    assert cents_a_euros(None) == ""


def test_euros_texto_a_cents_redondea_y_acepta_coma():
    assert euros_texto_a_cents("19.99") == 1999
    assert euros_texto_a_cents("19,99") == 1999
    assert euros_texto_a_cents("  20  ") == 2000


def test_euros_texto_a_cents_vacio_o_no_numerico_falla():
    with pytest.raises(ValueError):
        euros_texto_a_cents("")
    with pytest.raises(ValueError):
        euros_texto_a_cents("veinte euros")


# --------------------------------------------------------------------------
# Agregados
# --------------------------------------------------------------------------


def _fila_vendida(precio_final_cents: int, beneficio_cents: int) -> dict:
    return {
        "producto_id": "p1",
        "venta": {
            "precio_final_cents": precio_final_cents,
            "plataforma_venta": "wallapop",
            "estado": "vendida",
        },
        "beneficio_bruto_cents": beneficio_cents,
    }


def _fila_devuelta(precio_final_cents: int) -> dict:
    return {
        "producto_id": "p2",
        "venta": {
            "precio_final_cents": precio_final_cents,
            "plataforma_venta": "vinted",
            "estado": "devuelta",
        },
        # El store deja esto en None para todo lo que no sea 'vendida'.
        "beneficio_bruto_cents": None,
    }


def _fila_no_vendida() -> dict:
    return {"producto_id": "p3", "venta": None, "beneficio_bruto_cents": None}


def test_total_vendido_cents_solo_suma_estado_vendida():
    filas = [_fila_vendida(2000, 1500), _fila_devuelta(3000), _fila_no_vendida()]
    assert total_vendido_cents(filas) == 2000


def test_beneficio_total_cents_devuelta_no_suma():
    filas = [_fila_vendida(2000, 1500), _fila_vendida(1000, -200), _fila_devuelta(3000)]
    # 1500 + (-200) = 1300; la devuelta (beneficio=None) no aporta nada.
    assert beneficio_total_cents(filas) == 1300


def test_agregados_lista_vacia_son_cero():
    assert total_vendido_cents([]) == 0
    assert beneficio_total_cents([]) == 0


# --------------------------------------------------------------------------
# Export a Excel
# --------------------------------------------------------------------------


def _fila_completa() -> dict:
    return {
        "producto_id": "abc123",
        "lote_id": "lote1",
        "referencia": 7,
        "titulo": "Sudadera Reebok XXL",
        "coste_cents": 500,
        "publicaciones": [
            {
                "plataforma": "wallapop",
                "subido_en": "2026-07-20T10:00:00+00:00",
                "precio_elegido_cents": 2000,
                "tasacion": None,
            }
        ],
        "venta": {
            "precio_final_cents": 1800,
            "plataforma_venta": "wallapop",
            "estado": "vendida",
            "lote_venta_id": None,
            "fecha_venta": "2026-07-21T09:00:00+00:00",
        },
        "beneficio_bruto_cents": 1300,
    }


def _fila_no_vendida_completa() -> dict:
    return {
        "producto_id": "def456",
        "lote_id": "lote1",
        "referencia": 8,
        "titulo": "Camiseta Nike M",
        "coste_cents": 0,
        "publicaciones": [],
        "venta": None,
        "beneficio_bruto_cents": None,
    }


def test_exportar_excel_escribe_columnas_y_filas_legibles(tmp_path: Path):
    filas = [_fila_completa(), _fila_no_vendida_completa()]
    destino = tmp_path / "sub" / "finanzas.xlsx"

    ruta = exportar_excel(filas, destino)

    assert ruta == destino
    assert destino.exists()

    libro = openpyxl.load_workbook(destino)
    hoja = libro.active
    cabecera = [c.value for c in next(hoja.iter_rows(min_row=1, max_row=1))]
    assert tuple(cabecera) == _COLUMNAS_ESPERADAS

    filas_leidas = list(hoja.iter_rows(min_row=2, values_only=True))
    assert len(filas_leidas) == 2

    fila_vendida = filas_leidas[0]
    assert fila_vendida[0] == 7  # Referencia
    assert fila_vendida[1] == "Sudadera Reebok XXL"  # Título
    assert fila_vendida[3] == "5.00"  # Coste (€)
    assert "wallapop" in fila_vendida[4]  # Subido
    assert fila_vendida[5] == "sí"  # Vendido
    assert fila_vendida[6] == "18.00"  # Precio venta (€)
    assert fila_vendida[7] == "wallapop"  # Plataforma venta
    assert fila_vendida[9] == "13.00"  # Beneficio bruto (€)

    fila_sin_vender = filas_leidas[1]
    assert fila_sin_vender[5] == "no"
    # openpyxl no distingue "" de "sin valor": una cadena vacía escrita se
    # relee como `None` (celda vacía) -- ambas representan "no aplica".
    assert fila_sin_vender[6] in ("", None)
    assert fila_sin_vender[9] in ("", None)


def test_exportar_excel_devuelta_se_distingue_de_vendida(tmp_path: Path):
    fila = _fila_completa()
    fila["venta"]["estado"] = "devuelta"
    fila["beneficio_bruto_cents"] = None

    destino = tmp_path / "finanzas.xlsx"
    exportar_excel([fila], destino)

    libro = openpyxl.load_workbook(destino)
    hoja = libro.active
    fila_leida = next(hoja.iter_rows(min_row=2, max_row=2, values_only=True))
    assert fila_leida[5] == "devuelta"
    assert fila_leida[9] in ("", None)  # sin beneficio


def test_exportar_excel_crea_directorio_padre(tmp_path: Path):
    destino = tmp_path / "no_existe_todavia" / "finanzas.xlsx"
    assert not destino.parent.exists()
    exportar_excel([], destino)
    assert destino.exists()


def test_nombre_export_por_defecto_incluye_la_fecha():
    momento = datetime(2026, 7, 21, 12, 0, tzinfo=timezone.utc)
    assert nombre_export_por_defecto(momento) == "finanzas_2026-07-21.xlsx"


def test_export_xlsx_neutraliza_formulas(tmp_path):
    """Un titulo que empieza por `=` NO puede acabar como formula en el .xlsx.

    Clase "CSV/Excel injection": Excel y LibreOffice ejecutan cualquier celda de
    TEXTO que arranque por = + - @ (o tabulador/retorno). El titulo de una ficha
    es texto libre, asi que el vector es real. Riesgo bajo -- el fichero lo
    genera y lo abre Diego con sus propios datos -- pero el repo ya sanitiza el
    texto que va a las plataformas (`schema.validar_texto`) y que la hoja de
    calculo no tuviera nada era una incoherencia.

    Se comprueba sobre el fichero ESCRITO y releido, no sobre la funcion: la
    garantia es que el .xlsx no lleve la formula, no que exista un helper
    (`decision-making.md` §17).
    """
    from openpyxl import load_workbook

    filas = [
        {
            "referencia": 1,
            "titulo": "=1+1",
            "lote": "L",
            "coste_cents": 0,
            "publicaciones": [],
            "venta": None,
        },
        {
            "referencia": 2,
            "titulo": "Sudadera Reebok XXL",
            "lote": "L",
            "coste_cents": 0,
            "publicaciones": [],
            "venta": None,
        },
    ]
    destino = exportar_excel(filas, tmp_path / "f.xlsx")
    hoja = load_workbook(destino).active

    peligrosa = hoja.cell(row=2, column=2).value
    assert not str(peligrosa).startswith("="), f"formula viva en la celda: {peligrosa!r}"
    assert "1+1" in str(peligrosa), "el valor debe seguir siendo LEGIBLE, no borrado"
    # Un titulo normal no se toca.
    assert hoja.cell(row=3, column=2).value == "Sudadera Reebok XXL"
